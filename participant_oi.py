#!/usr/bin/env python3
"""
NSE Participant-wise Open Interest -> Excel report.

Fetches the "F&O - Participant wise Open Interest" report from NSE for the 5 most
recent trading days and builds an Excel workbook mirroring the analyst layout:

  * Left block   "Futures & Options"           - day-over-day change per participant
                                                  (Added/Closed Longs & Shorts, Net Buy/Sell)
  * Middle block "Total Positions Carried"      - net position (Long-Short) for
                                                  TODAY back through 4 DAYS AGO
  * Right block  "Positions Bought / Sold Today"- the net-change data pivoted by participant

A compact "India VIX" strip below the middle block tracks the closing VIX over the
same 5 sessions, followed by a "Nifty SRT" strip (Speculation Ratio Territory:
Nifty close / 124-day SMA).

Raw daily CSVs are cached under data/ so historical days are never re-fetched.
Run any time after ~7 PM IST; NSE publishes the file after market close.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import sys
import urllib.parse
from datetime import date, datetime, timedelta
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from zoneinfo import ZoneInfo

    IST = ZoneInfo("Asia/Kolkata")
except Exception:  # pragma: no cover - fallback if tz data is unavailable
    from datetime import timezone

    IST = timezone(timedelta(hours=5, minutes=30))

# --------------------------------------------------------------------------- #
# Config
# --------------------------------------------------------------------------- #
ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
OUTPUT_FILE = ROOT / "participant_oi.xlsx"
HASH_FILE = ROOT / "report_data.hash"  # fingerprint of source data; drives commit-on-change

URL_TEMPLATE = "https://archives.nseindia.com/content/nsccl/fao_participant_oi_{ddmmyyyy}.csv"
INDEX_URL_TEMPLATE = "https://archives.nseindia.com/content/indices/ind_close_all_{ddmmyyyy}.csv"
NSE_HOME = "https://www.nseindia.com/"
NSE_OC_PAGE = "https://www.nseindia.com/option-chain"
OC_CONTRACT_URL = "https://www.nseindia.com/api/option-chain-contract-info?symbol={sym}"
OC_V3_URL = "https://www.nseindia.com/api/option-chain-v3?type=Indices&symbol={sym}&expiry={exp}"
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
)
MAX_LOOKBACK_DAYS = 15  # safety bound when walking back over holidays
NUM_DAYS = 5            # today .. 4 days ago (net-positions-carried horizon)
SRT_SMA_DAYS = 124        # ~half a trading year; NK StockTalk SRT convention


def _day_label(i: int) -> str:
    """Column heading for the i-th most recent session (0 = today)."""
    if i == 0:
        return "TODAY"
    return f"{i} DAY{'S' if i > 1 else ''} AGO"


DAY_LABELS = [_day_label(i) for i in range(NUM_DAYS)]

# Column layout (1-based). Left block = cols 1-7 (col 8 spacer). The middle
# "Total Positions Carried" block is a label column plus one column per day; the
# right block base column follows a one-column gutter, so it shifts automatically
# with NUM_DAYS instead of being hardcoded.
MID_LABEL_COL = 9
MID_DAY_COL0 = 10                              # first (TODAY) day column
SPACER_MID_RIGHT = MID_DAY_COL0 + NUM_DAYS     # blank gutter after the day columns
RB = SPACER_MID_RIGHT + 1                      # right-block base column

# Participant rows: (CSV key, display label)
PARTICIPANTS = [
    ("Client", "Clients"),
    ("DII", "DIIs"),
    ("FII", "FIIs"),
    ("Pro", "Pro"),
]

# CSV column layout (0-based), header line 2 of the file.
COLS = {
    "Future Index Long": 1,
    "Future Index Short": 2,
    "Future Stock Long": 3,
    "Future Stock Short": 4,
    "Option Index Call Long": 5,
    "Option Index Put Long": 6,
    "Option Index Call Short": 7,
    "Option Index Put Short": 8,
    "Option Stock Call Long": 9,
    "Option Stock Put Long": 10,
    "Option Stock Call Short": 11,
    "Option Stock Put Short": 12,
}

# Instruments: (display name, long-col, short-col, long_is_bullish)
# long_is_bullish: a long future or long call is bullish, but a long PUT is bearish,
# so puts invert the sentiment colouring of every long/short/net cell.
INSTRUMENTS = [
    ("Index Futures", "Future Index Long", "Future Index Short", True),
    ("Index Call", "Option Index Call Long", "Option Index Call Short", True),
    ("Index Put", "Option Index Put Long", "Option Index Put Short", False),
    ("Stock Futures", "Future Stock Long", "Future Stock Short", True),
    ("Stock Calls", "Option Stock Call Long", "Option Stock Call Short", True),
    ("Stock Puts", "Option Stock Put Long", "Option Stock Put Short", False),
]

# --------------------------------------------------------------------------- #
# Fetch + parse
# --------------------------------------------------------------------------- #
def _fetch_raw(d: date) -> str | None:
    """Return the CSV text for date d, or None if NSE has no file (404 = holiday)."""
    cache = DATA_DIR / f"fao_participant_oi_{d.isoformat()}.csv"
    if cache.exists():
        return cache.read_text()

    url = URL_TEMPLATE.format(ddmmyyyy=d.strftime("%d%m%Y"))
    try:
        resp = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=30)
    except requests.RequestException as exc:
        print(f"  ! network error for {d}: {exc}", file=sys.stderr)
        return None

    if resp.status_code == 404:
        return None
    resp.raise_for_status()

    text = resp.text
    if "Participant wise Open Interest" not in text:
        return None

    DATA_DIR.mkdir(exist_ok=True)
    cache.write_text(text)
    return text


def _parse(text: str) -> dict[str, dict[str, int]]:
    """Parse CSV text into {participant: {column-name: value}}."""
    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader if r]
    out: dict[str, dict[str, int]] = {}
    for row in rows:
        key = row[0].strip()
        if key not in ("Client", "DII", "FII", "Pro", "TOTAL"):
            continue
        out[key] = {name: int(row[idx].strip() or 0) for name, idx in COLS.items()}
    return out


def load_recent_days(n: int = NUM_DAYS) -> list[tuple[date, dict]]:
    """Return the n most recent trading days as (date, parsed) newest first."""
    days: list[tuple[date, dict]] = []
    d = datetime.now(IST).date()
    steps = 0
    while len(days) < n and steps < MAX_LOOKBACK_DAYS:
        text = _fetch_raw(d)
        if text:
            parsed = _parse(text)
            if parsed:
                print(f"  loaded {d}")
                days.append((d, parsed))
        d -= timedelta(days=1)
        steps += 1
    if len(days) < n:
        raise SystemExit(
            f"Only found {len(days)} trading day(s) in the last {MAX_LOOKBACK_DAYS} days; "
            f"need {n}. NSE may not have published today's file yet."
        )
    return days


def net(day: dict, participant: str, long_col: str, short_col: str) -> int:
    """Net position = Long - Short for a participant/instrument on a given day."""
    p = day[participant]
    return p[long_col] - p[short_col]


def _fetch_index_close_csv(d: date) -> str | None:
    """Return the ind_close_all CSV text for date d (cached), or None if unavailable."""
    cache = DATA_DIR / f"ind_close_all_{d.isoformat()}.csv"
    if cache.exists():
        return cache.read_text()
    url = INDEX_URL_TEMPLATE.format(ddmmyyyy=d.strftime("%d%m%Y"))
    try:
        resp = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=30)
    except requests.RequestException:
        return None
    if resp.status_code == 200 and "Index Name" in resp.text:
        DATA_DIR.mkdir(exist_ok=True)
        cache.write_text(resp.text)
        return resp.text
    return None


def _index_row(text: str, name: str) -> list[str] | None:
    """First row of an index-close CSV whose Index Name matches `name`."""
    for row in csv.reader(io.StringIO(text)):
        if row and row[0].strip() == name:
            return row
    return None


def fetch_nifty_ohlc(d: date, max_back: int = 7):
    """Return (date, open, high, low, close) for Nifty 50 on/just before d, or None."""
    cur = d
    for _ in range(max_back):
        text = _fetch_index_close_csv(cur)
        if text:
            row = _index_row(text, "Nifty 50")
            if row:
                return (cur, float(row[2]), float(row[3]), float(row[4]), float(row[5]))
        cur -= timedelta(days=1)
    return None


def fetch_vix_history(days: list[tuple[date, dict]]) -> list[tuple[date, float | None]]:
    """Closing India VIX for each report day, aligned by date (None if unavailable).

    India VIX ships in the same ind_close_all file as the Nifty OHLC (closing value
    is column index 5), so this reuses the cached per-day download.
    """
    out: list[tuple[date, float | None]] = []
    for d, _ in days:
        text = _fetch_index_close_csv(d)
        row = _index_row(text, "India VIX") if text else None
        out.append((d, float(row[5]) if row else None))
    return out


def fetch_nifty_close_history(end: date, trading_days: int) -> list[tuple[date, float]]:
    """Oldest-first Nifty 50 closes over `trading_days` sessions ending at `end`."""
    collected: list[tuple[date, float]] = []
    cur = end
    steps = 0
    max_steps = trading_days * 2 + 60  # weekends and holiday buffer
    while len(collected) < trading_days and steps < max_steps:
        text = _fetch_index_close_csv(cur)
        if text:
            row = _index_row(text, "Nifty 50")
            if row:
                collected.append((cur, float(row[5])))
        cur -= timedelta(days=1)
        steps += 1
    collected.reverse()
    return collected


def _srt_for_date(closes: list[tuple[date, float]], as_of: date):
    """Return (close, sma124, srt) for `as_of`, or (None, None, None) if unavailable."""
    by_date = {d: c for d, c in closes}
    if as_of not in by_date:
        return None, None, None
    eligible = [(d, c) for d, c in closes if d <= as_of]
    if len(eligible) < SRT_SMA_DAYS:
        return None, None, None
    window = eligible[-SRT_SMA_DAYS:]
    spot = by_date[as_of]
    sma = sum(c for _, c in window) / SRT_SMA_DAYS
    return spot, sma, spot / sma


def fetch_srt_history(days: list[tuple[date, dict]]):
    """Nifty SRT inputs for each report day: (date, close, sma124, srt_ratio)."""
    newest = days[0][0]
    needed = SRT_SMA_DAYS + NUM_DAYS + 30
    closes = fetch_nifty_close_history(newest, needed)
    return [(_d, *_srt_for_date(closes, _d)) for _d, _ in days]


def _srt_font(value: float | None):
    """Colour SRT by the usual NK StockTalk accumulation / exit bands."""
    if value is None:
        return None
    if value <= 0.9:
        return GREEN_B
    if value >= 1.3:
        return RED_B
    return BOLD


def pivot_levels(high: float, low: float, close: float) -> dict[str, float]:
    """Classic floor-trader pivot points + Central Pivot Range for the next session."""
    pp = (high + low + close) / 3
    return {
        "pp": pp,
        "r1": 2 * pp - low,
        "s1": 2 * pp - high,
        "r2": pp + (high - low),
        "s2": pp - (high - low),
        "r3": high + 2 * (pp - low),
        "s3": low - 2 * (high - pp),
        "bc": (high + low) / 2,
        "tc": 2 * pp - (high + low) / 2,
    }


def fetch_nifty_option_chain(symbol: str = "NIFTY") -> dict | None:
    """Live option chain for the nearest (this-week) NIFTY expiry, or None on failure.

    Returns {expiry, spot, pcr, max_pain, strikes:[{strike, ce_oi, ce_chg, ce_vol,
    pe_oi, pe_chg, pe_vol}]}. OI/volume are as published by NSE (end-of-day after close).
    """
    try:
        s = requests.Session()
        s.headers.update({
            "User-Agent": USER_AGENT,
            "Accept-Language": "en-US,en;q=0.9",
            "Sec-Fetch-Dest": "document", "Sec-Fetch-Mode": "navigate", "Sec-Fetch-Site": "none",
        })
        s.get(NSE_HOME, timeout=30)        # seed Akamai cookies
        s.get(NSE_OC_PAGE, timeout=30)
        api_headers = {
            "Accept": "application/json, text/plain, */*",
            "Referer": NSE_OC_PAGE,
            "Sec-Fetch-Dest": "empty", "Sec-Fetch-Mode": "cors", "Sec-Fetch-Site": "same-origin",
        }
        info = s.get(OC_CONTRACT_URL.format(sym=symbol), headers=api_headers, timeout=30).json()
        expiry = info["expiryDates"][0]  # nearest upcoming expiry = this week
        url = OC_V3_URL.format(sym=symbol, exp=urllib.parse.quote(expiry))
        data = s.get(url, headers=api_headers, timeout=30).json()
    except Exception as exc:  # network / bot-block / shape change -> skip gracefully
        print(f"  ! option chain unavailable: {type(exc).__name__}: {exc}", file=sys.stderr)
        return None

    rec = data.get("records", {})
    spot = rec.get("underlyingValue")
    strikes = []
    for q in rec.get("data", []):
        ce, pe = q.get("CE"), q.get("PE")
        strikes.append({
            "strike": q["strikePrice"],
            "ce_oi": ce["openInterest"] if ce else 0,
            "ce_chg": ce["changeinOpenInterest"] if ce else 0,
            "ce_vol": ce["totalTradedVolume"] if ce else 0,
            "pe_oi": pe["openInterest"] if pe else 0,
            "pe_chg": pe["changeinOpenInterest"] if pe else 0,
            "pe_vol": pe["totalTradedVolume"] if pe else 0,
        })
    if not strikes or spot is None:
        return None

    tot_ce = sum(r["ce_oi"] for r in strikes)
    tot_pe = sum(r["pe_oi"] for r in strikes)
    pcr = (tot_pe / tot_ce) if tot_ce else 0.0

    # Max pain = expiry strike minimising total payout by option writers.
    def payout(k):
        return (sum(r["ce_oi"] * max(0, k - r["strike"]) for r in strikes)
                + sum(r["pe_oi"] * max(0, r["strike"] - k) for r in strikes))
    max_pain = min((r["strike"] for r in strikes), key=payout)

    return {"expiry": expiry, "spot": spot, "pcr": pcr, "max_pain": max_pain, "strikes": strikes}


def _strength(vol: int, mx: int) -> str:
    """Grade a level's conviction from its traded volume relative to the strongest shown."""
    f = (vol / mx) if mx else 0.0
    return "High" if f >= 0.66 else "Medium" if f >= 0.33 else "Low"


def analyze_oi_sr(oc: dict, n: int = 3) -> dict:
    """Top-n Call-OI resistance walls and Put-OI support walls, graded by volume."""
    spot, strikes = oc["spot"], oc["strikes"]
    atm = min(strikes, key=lambda r: abs(r["strike"] - spot))["strike"]
    res = sorted([r for r in strikes if r["strike"] >= atm and r["ce_oi"] > 0],
                 key=lambda r: r["ce_oi"], reverse=True)[:n]
    sup = sorted([r for r in strikes if r["strike"] <= atm and r["pe_oi"] > 0],
                 key=lambda r: r["pe_oi"], reverse=True)[:n]
    res.sort(key=lambda r: r["strike"])              # nearest resistance above spot first
    sup.sort(key=lambda r: r["strike"], reverse=True)  # nearest support below spot first
    mx = max([r["ce_vol"] for r in res] + [r["pe_vol"] for r in sup] + [1])
    return {"atm": atm, "resistance": res, "support": sup, "max_vol": mx}


# --------------------------------------------------------------------------- #
# Excel styling
# --------------------------------------------------------------------------- #
NAVY = PatternFill("solid", fgColor="1F2A44")
SUBHEAD = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
INSTR_FILL = PatternFill("solid", fgColor="EDEDED")

WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
GREEN = Font(color="0B7A34")
RED = Font(color="C00000")
GREEN_B = Font(color="0B7A34", bold=True)
RED_B = Font(color="C00000", bold=True)
GREY = Font(color="808080", italic=True)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

IND_FMT = "#,##,##0;-#,##,##0;0"  # Indian digit grouping, kept numeric
NIFTY_FMT = "#,##,##0.00"          # index level with two decimals
VIX_FMT = "0.00"                   # volatility index, two decimals
SRT_FMT = "0.000"                  # speculation ratio, three decimals


def _put(ws, r, c, value, *, font=None, fill=None, align=None, fmt=None, border=True):
    cell = ws.cell(row=r, column=c, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = BORDER
    return cell


def _colour(bullish):
    return GREEN if bullish else RED


def _long_label(delta, long_bullish=True):
    # Adding to a bullish long is bullish; for puts (long_bullish=False) it inverts.
    label = "Added Longs" if delta >= 0 else "Closed Longs"
    bullish = (delta >= 0) if long_bullish else (delta < 0)
    return label, _colour(bullish)


def _short_label(delta, long_bullish=True):
    # A short is the opposite stance of the long, so its sentiment is inverted again.
    label = "Added Shorts" if delta >= 0 else "Closed Shorts"
    bullish = (delta < 0) if long_bullish else (delta >= 0)
    return label, _colour(bullish)


def _net_label(delta, long_bullish=True):
    # Net buying is bullish for futures/calls, bearish for puts (buying puts = bearish).
    label = "Bought Net" if delta >= 0 else "Sold Net"
    bullish = (delta >= 0) if long_bullish else (delta < 0)
    return label, _colour(bullish)


# --------------------------------------------------------------------------- #
# Build workbook
# --------------------------------------------------------------------------- #
def build(days: list[tuple[date, dict]], ohlc, oc: dict | None,
          vix_hist: list[tuple[date, float | None]],
          srt_hist: list[tuple[date, float | None, float | None, float | None]],
          path: Path) -> None:
    # Day-over-day change (left & right blocks) compares today vs 1 day ago; the
    # middle block spans all loaded days.
    (dt_today, day_today) = days[0]
    (dt_1, day_1) = days[1]

    wb = Workbook()
    ws = wb.active
    ws.title = "Participant OI"
    ws.sheet_view.showGridLines = False

    # ---- title row -------------------------------------------------------- #
    _put(ws, 1, 1, "Futures & Options", font=WHITE_BOLD, fill=NAVY, align=LEFT)
    for c in range(2, 8):
        _put(ws, 1, c, None, fill=NAVY)
    _put(ws, 1, MID_LABEL_COL, "Total Positions Carried", font=WHITE_BOLD, fill=NAVY, align=CENTER)
    for c in range(MID_LABEL_COL + 1, MID_LABEL_COL + 1 + NUM_DAYS):
        _put(ws, 1, c, None, fill=NAVY)
    _put(ws, 1, RB, "Positions Bought / Sold Today", font=WHITE_BOLD, fill=NAVY, align=CENTER)
    for c in (RB + 1, RB + 2, RB + 3):
        _put(ws, 1, c, None, fill=NAVY)

    r = 2
    # Each instrument section occupies: 1 sub-header + 4 participants + 1 total = 6 rows
    for instr_name, long_col, short_col, long_bullish in INSTRUMENTS:
        # ---- LEFT block sub-headers ---- #
        _put(ws, r, 1, "", fill=INSTR_FILL)
        _put(ws, r, 2, f"{instr_name} Longs", font=BOLD, fill=SUBHEAD, align=CENTER)
        _put(ws, r, 3, "", fill=SUBHEAD)
        _put(ws, r, 4, f"{instr_name} Shorts", font=BOLD, fill=SUBHEAD, align=CENTER)
        _put(ws, r, 5, "", fill=SUBHEAD)
        _put(ws, r, 6, "Net Buy / Sell for Today", font=BOLD, fill=SUBHEAD, align=CENTER)
        _put(ws, r, 7, "", fill=SUBHEAD)
        # ---- MIDDLE block sub-headers ---- #
        _put(ws, r, MID_LABEL_COL, instr_name, font=BOLD, fill=SUBHEAD, align=LEFT)
        for i, label in enumerate(DAY_LABELS):
            _put(ws, r, MID_DAY_COL0 + i, label, font=BOLD, fill=SUBHEAD, align=CENTER)
        r += 1

        tot_dlong = tot_dshort = 0
        for csv_key, disp in PARTICIPANTS:
            long_t = day_today[csv_key][long_col]
            short_t = day_today[csv_key][short_col]
            long_y = day_1[csv_key][long_col]
            short_y = day_1[csv_key][short_col]
            d_long = long_t - long_y
            d_short = short_t - short_y
            d_net = d_long - d_short
            tot_dlong += d_long
            tot_dshort += d_short

            l_lbl, l_font = _long_label(d_long, long_bullish)
            s_lbl, s_font = _short_label(d_short, long_bullish)
            n_lbl, n_font = _net_label(d_net, long_bullish)

            _put(ws, r, 1, disp, font=BOLD, align=LEFT)
            _put(ws, r, 2, l_lbl, font=l_font, align=LEFT)
            _put(ws, r, 3, d_long, font=l_font, align=RIGHT, fmt=IND_FMT)
            _put(ws, r, 4, s_lbl, font=s_font, align=LEFT)
            _put(ws, r, 5, d_short, font=s_font, align=RIGHT, fmt=IND_FMT)
            _put(ws, r, 6, n_lbl, font=n_font, align=LEFT)
            _put(ws, r, 7, d_net, font=n_font, align=RIGHT, fmt=IND_FMT)

            # ---- MIDDLE block: net positions carried, TODAY .. N days ago ---- #
            _put(ws, r, MID_LABEL_COL, disp, font=BOLD, align=LEFT)
            for i, (_dt_i, day_i) in enumerate(days):
                _put(ws, r, MID_DAY_COL0 + i,
                     net(day_i, csv_key, long_col, short_col), align=RIGHT, fmt=IND_FMT)
            r += 1

        # ---- Total row ---- #
        tot_dnet = tot_dlong - tot_dshort
        _put(ws, r, 1, "Total", font=BOLD, fill=TOTAL_FILL, align=LEFT)
        _put(ws, r, 2, "", fill=TOTAL_FILL)
        _put(ws, r, 3, tot_dlong, font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=IND_FMT)
        _put(ws, r, 4, "", fill=TOTAL_FILL)
        _put(ws, r, 5, tot_dshort, font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=IND_FMT)
        _put(ws, r, 6, "", fill=TOTAL_FILL)
        _put(ws, r, 7, tot_dnet, font=BOLD, fill=TOTAL_FILL, align=RIGHT, fmt=IND_FMT)
        for c in range(MID_LABEL_COL, MID_LABEL_COL + 1 + NUM_DAYS):
            _put(ws, r, c, "Total" if c == MID_LABEL_COL else "",
                 font=BOLD, fill=TOTAL_FILL, align=LEFT)
        r += 1

    # ---- India VIX close over the same sessions (mirrors the day columns) ---- #
    if vix_hist:
        vr = r + 1  # one blank row below the instrument tables
        _put(ws, vr, MID_LABEL_COL, "India VIX", font=WHITE_BOLD, fill=NAVY, align=LEFT)
        for i, label in enumerate(DAY_LABELS):
            _put(ws, vr, MID_DAY_COL0 + i, label, font=WHITE_BOLD, fill=NAVY, align=CENTER)
        _put(ws, vr + 1, MID_LABEL_COL, "Close", font=BOLD, fill=SUBHEAD, align=LEFT)
        for i, (_d, v) in enumerate(vix_hist):
            _put(ws, vr + 1, MID_DAY_COL0 + i,
                 v if v is not None else "-",
                 align=RIGHT, fmt=(VIX_FMT if v is not None else None))
        r = vr + 2

    # ---- Nifty SRT (Speculation Ratio Territory) below VIX ----------------- #
    if srt_hist:
        sr = r
        _put(ws, sr, MID_LABEL_COL, "Nifty SRT", font=WHITE_BOLD, fill=NAVY, align=LEFT)
        for i, label in enumerate(DAY_LABELS):
            _put(ws, sr, MID_DAY_COL0 + i, label, font=WHITE_BOLD, fill=NAVY, align=CENTER)
        _put(ws, sr + 1, MID_LABEL_COL, "SRT", font=BOLD, fill=SUBHEAD, align=LEFT)
        for i, (_d, _spot, _sma, srt) in enumerate(srt_hist):
            _put(ws, sr + 1, MID_DAY_COL0 + i,
                 srt if srt is not None else "-",
                 font=_srt_font(srt), align=RIGHT, fmt=(SRT_FMT if srt is not None else None))
        _put(ws, sr + 2, MID_LABEL_COL, f"{SRT_SMA_DAYS}-day SMA", font=BOLD, fill=SUBHEAD, align=LEFT)
        for i, (_d, _spot, sma, _srt) in enumerate(srt_hist):
            _put(ws, sr + 2, MID_DAY_COL0 + i,
                 sma if sma is not None else "-",
                 align=RIGHT, fmt=(NIFTY_FMT if sma is not None else None))
        _put(ws, sr + 3, MID_LABEL_COL, "Nifty Close", font=BOLD, fill=SUBHEAD, align=LEFT)
        for i, (_d, spot, _sma, _srt) in enumerate(srt_hist):
            _put(ws, sr + 3, MID_DAY_COL0 + i,
                 spot if spot is not None else "-",
                 align=RIGHT, fmt=(NIFTY_FMT if spot is not None else None))
        r = sr + 4

    # ---- RIGHT block: Positions Bought / Sold Today, grouped by participant --- #
    rr = 2
    for csv_key, disp in PARTICIPANTS:
        for instr_name, long_col, short_col, long_bullish in INSTRUMENTS:
            d_net = (
                (day_today[csv_key][long_col] - day_1[csv_key][long_col])
                - (day_today[csv_key][short_col] - day_1[csv_key][short_col])
            )
            n_lbl, n_font = _net_label(d_net, long_bullish)
            _put(ws, rr, RB, disp, font=BOLD, align=LEFT)
            _put(ws, rr, RB + 1, n_lbl, font=n_font, align=LEFT)
            _put(ws, rr, RB + 2, instr_name, align=LEFT)
            _put(ws, rr, RB + 3, d_net, font=n_font, align=RIGHT, fmt=IND_FMT)
            rr += 1

    # ---- Nifty support & resistance (bottom-right, below the right block) ---- #
    sr_last = rr
    if ohlc:
        ndate, o, h, l, c = ohlc
        p = pivot_levels(h, l, c)
        s = rr + 2  # leave a blank row under the right block

        _put(ws, s, RB, "Nifty 50  -  Support & Resistance", font=WHITE_BOLD, fill=NAVY, align=CENTER)
        for cc in (RB + 1, RB + 2, RB + 3):
            _put(ws, s, cc, None, fill=NAVY)
        _put(ws, s + 1, RB, f"Next session  -  based on {ndate:%d-%b-%Y} OHLC",
             font=BOLD, fill=SUBHEAD, align=CENTER)
        for cc in (RB + 1, RB + 2, RB + 3):
            _put(ws, s + 1, cc, "", fill=SUBHEAD)

        # OHLC line
        _put(ws, s + 2, RB, "Open", font=BOLD, align=LEFT)
        _put(ws, s + 2, RB + 1, o, align=RIGHT, fmt=NIFTY_FMT)
        _put(ws, s + 2, RB + 2, "High", font=BOLD, align=LEFT)
        _put(ws, s + 2, RB + 3, h, align=RIGHT, fmt=NIFTY_FMT)
        _put(ws, s + 3, RB, "Low", font=BOLD, align=LEFT)
        _put(ws, s + 3, RB + 1, l, align=RIGHT, fmt=NIFTY_FMT)
        _put(ws, s + 3, RB + 2, "Close", font=BOLD, align=LEFT)
        _put(ws, s + 3, RB + 3, c, align=RIGHT, fmt=NIFTY_FMT)

        # Resistances / pivot / supports (col 14-15) beside CPR (col 16-17)
        levels = [
            ("R3", p["r3"], RED_B), ("R2", p["r2"], RED_B), ("R1", p["r1"], RED_B),
            ("Pivot", p["pp"], BOLD),
            ("S1", p["s1"], GREEN_B), ("S2", p["s2"], GREEN_B), ("S3", p["s3"], GREEN_B),
        ]
        cpr = [("CPR Top", p["tc"], BOLD), ("Pivot", p["pp"], BOLD), ("CPR Bottom", p["bc"], BOLD)]
        for i, (name, val, font) in enumerate(levels):
            row = s + 4 + i
            _put(ws, row, RB, name, font=font, align=LEFT)
            _put(ws, row, RB + 1, val, font=font, align=RIGHT, fmt=NIFTY_FMT)
            if i < len(cpr):
                cname, cval, cfont = cpr[i]
                _put(ws, row, RB + 2, cname, font=cfont, align=LEFT)
                _put(ws, row, RB + 3, cval, font=cfont, align=RIGHT, fmt=NIFTY_FMT)
            else:
                _put(ws, row, RB + 2, "", border=False)
                _put(ws, row, RB + 3, "", border=False)
        sr_last = s + 4 + len(levels)

    # ---- Nifty option-chain OI support & resistance, graded by volume ---- #
    if oc:
        a = analyze_oi_sr(oc)
        mx = a["max_vol"]
        o = sr_last + 2

        def _merge(row, text, font, fill):
            ws.merge_cells(start_row=row, start_column=RB, end_row=row, end_column=RB + 4)
            _put(ws, row, RB, text, font=font, fill=fill, align=CENTER)
            for cc in range(RB + 1, RB + 5):
                _put(ws, row, cc, None, fill=fill)

        _merge(o, f"Nifty Option-Chain S/R   -   Weekly expiry {oc['expiry']}", WHITE_BOLD, NAVY)
        _merge(o + 1,
               f"Spot {oc['spot']:,.0f}    PCR(OI) {oc['pcr']:.2f}    Max Pain {oc['max_pain']:,.0f}",
               BOLD, SUBHEAD)

        def _headers(row, oi_label):
            for cc, txt in zip(range(RB, RB + 5), ["Strike", oi_label, "Chg OI", "Volume", "Strength"]):
                _put(ws, row, cc, txt, font=BOLD, fill=INSTR_FILL, align=CENTER)

        def _level_row(row, rec, oi, chg, vol, strike_font):
            lbl = _strength(vol, mx)
            sfont = BOLD if lbl == "High" else (GREY if lbl == "Low" else None)
            _put(ws, row, RB, rec["strike"], font=strike_font, align=RIGHT, fmt="#,##0")
            _put(ws, row, RB + 1, oi, align=RIGHT, fmt=IND_FMT)
            _put(ws, row, RB + 2, chg, font=(GREEN if chg >= 0 else RED), align=RIGHT, fmt=IND_FMT)
            _put(ws, row, RB + 3, vol, align=RIGHT, fmt=IND_FMT)
            _put(ws, row, RB + 4, lbl, font=sfont, align=CENTER)

        row = o + 2
        _merge(row, "RESISTANCE   -   Call OI walls (higher = stronger cap)", RED_B, SUBHEAD)
        _headers(row + 1, "Call OI")
        row += 2
        for rec in a["resistance"]:
            _level_row(row, rec, rec["ce_oi"], rec["ce_chg"], rec["ce_vol"], RED_B)
            row += 1

        _merge(row, "SUPPORT   -   Put OI walls (higher = stronger floor)", GREEN_B, SUBHEAD)
        _headers(row + 1, "Put OI")
        row += 2
        for rec in a["support"]:
            _level_row(row, rec, rec["pe_oi"], rec["pe_chg"], rec["pe_vol"], GREEN_B)
            row += 1
        sr_last = row

    # ---- footer with the dates used ---- #
    foot = max(r, rr, sr_last) + 1
    dates_str = "    ".join(f"{label} = {d:%d-%b-%Y}" for label, (d, _) in zip(DAY_LABELS, days))
    _put(
        ws, foot, 1,
        f"{dates_str}    (generated {datetime.now(IST):%d-%b-%Y %H:%M IST})",
        font=Font(italic=True, color="808080"), border=False,
    )

    # ---- column widths ---- #
    widths = {1: 9, 2: 13, 3: 11, 4: 13, 5: 11, 6: 16, 7: 11, 8: 2, MID_LABEL_COL: 15}
    for i in range(NUM_DAYS):                       # middle-block day columns
        widths[MID_DAY_COL0 + i] = 12
    widths[SPACER_MID_RIGHT] = 2                    # gutter before the right block
    widths.update({RB: 10, RB + 1: 12, RB + 2: 15, RB + 3: 12, RB + 4: 10})
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    wb.save(path)


def data_fingerprint(days: list[tuple[date, dict]], ohlc, vix_hist, srt_hist) -> str:
    """Deterministic hash of the report's source data (ignores generation time).

    Covers the loaded days of participant OI, the Nifty OHLC, the India VIX
    history, and the Nifty SRT inputs - the archive-based, date-stamped inputs.
    Unchanged inputs -> identical hash -> no new commit.
    """
    payload = {
        "days": [[d.isoformat(), day] for d, day in days],
        "ohlc": [ohlc[0].isoformat(), *ohlc[1:]] if ohlc else None,
        "vix": [[d.isoformat(), v] for d, v in vix_hist] if vix_hist else None,
        "srt": [[d.isoformat(), spot, sma, srt] for d, spot, sma, srt in srt_hist] if srt_hist else None,
    }
    blob = json.dumps(payload, sort_keys=True, default=str)
    return hashlib.sha256(blob.encode()).hexdigest()


def main() -> None:
    today = datetime.now(IST).date()

    # Skip weekends & market holidays: NSE publishes no participant file that day.
    if _fetch_raw(today) is None:
        print(f"{today:%d-%b-%Y}: weekend/market holiday - no NSE data published, skipping run.")
        return

    print("Fetching NSE participant-wise OI...")
    days = load_recent_days()
    ohlc = fetch_nifty_ohlc(days[0][0])
    oc = fetch_nifty_option_chain("NIFTY")
    vix_hist = fetch_vix_history(days)
    srt_hist = fetch_srt_history(days)
    print(f"Building {OUTPUT_FILE.name}...")
    build(days, ohlc, oc, vix_hist, srt_hist, OUTPUT_FILE)
    HASH_FILE.write_text(data_fingerprint(days, ohlc, vix_hist, srt_hist) + "\n")
    print(f"Done -> {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
